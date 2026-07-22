import json

import pytest

from evaluation.benchmark import load_records
from evaluation.excel_report import generate_workbook
from evaluation.quality_gates import assess


def valid_record():
    source = {"document_id": "paper_ms", "page": 3, "chunk_id": "ms_1"}
    return {"id": "B1", "question": "Define a compiler.", "category": "DEFINITION", "difficulty": "BEGINNER",
            "expected_intent": "EXAM_ANSWER", "gold_answer": "Translates a whole program.",
            "ground_truth_references": [source], "expected_pages": [3], "expected_topics": ["compiler"],
            "expected_citations": [source], "review_status": "AUTO_VALIDATED_EXACT",
            "generation_provenance": "exact_question_mark_scheme_pair", "question_chunk_id": "qp_1", "answer_chunk_id": "ms_1"}


def write(path, row):
    path.write_text(json.dumps(row) + "\n", encoding="utf-8")


def test_benchmark_accepts_only_defensible_exact_pairs(tmp_path):
    path = tmp_path / "benchmark.jsonl"
    write(path, valid_record())
    assert load_records(path)[0]["id"] == "B1"


@pytest.mark.parametrize("change", [{"gold_answer": ""}, {"question": "Broken Ã text"}, {"ground_truth_references": []}])
def test_benchmark_rejects_empty_corrupt_or_unreferenced_gold(tmp_path, change):
    path, row = tmp_path / "bad.jsonl", valid_record()
    row.update(change)
    write(path, row)
    with pytest.raises(ValueError):
        load_records(path)


def test_quality_gates_cannot_pass_unmeasured_semantics():
    report = {"benchmark": {"measurable_count": 120}, "ingestion": {"coverage": 1},
              "retrieval": {"coverage": 1, "recall_at_5": 1, "recall_at_10": 1, "mrr": 1, "ndcg_at_10": 1, "exact_scheme_retrieval_accuracy": 1},
              "answers": {"coverage": 1, "citation_identity_accuracy": 1, "citation_coverage": 1, "technical_failure_rate": 0},
              "performance": {"median_latency_ms": 1, "p95_latency_ms": 2},
              "ragas": {"status": "NOT_MEASURED", "reason": "not run"}}
    gates = assess(report)
    assert gates["all_measured_gates_passed"]
    assert not gates["all_required_gates_passed"]


def test_excel_report_retains_failed_and_missing_answers(tmp_path):
    first, second = valid_record(), {**valid_record(), "id": "B2", "question": "Second question"}
    report = {
        "retrieval": {"rows": [{"id": "B1", "execution_status": "COMPLETED"}, {"id": "B2", "execution_status": "FAILED", "error": "retrieval failed"}]},
        "answers": {"rows": [{"id": "B1", "execution_status": "COMPLETED", "answer": "Answer"}]},
        "ragas": {"rows": []}, "failure_analysis": {"rows": [{"id": "B2", "failure_category": "RETRIEVAL_EXECUTION_FAILURE"}]},
        "benchmark": {"measurable_count": 2}, "quality_gates": {"gates": {}},
        "generated_at": "now", "dataset": "test", "evaluation_limit": None, "pilot_run": True, "system_versions": {},
    }
    path = generate_workbook(report, [first, second], tmp_path / "report.xlsx")
    from openpyxl import load_workbook
    sheet = load_workbook(path, read_only=True)["Question Results"]
    assert sheet.max_row == 3
    assert sheet.cell(3, 5).value == "FAILED"
    assert sheet.cell(3, 7).value == "retrieval failed"
