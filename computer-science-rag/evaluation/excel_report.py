"""Excel evidence report with explicit execution and measurement statuses."""
from __future__ import annotations

import json
from pathlib import Path


COLUMNS = [
    "Question ID", "Question", "Category", "Difficulty", "Execution Status", "Row Result", "Execution Error",
    "Retrieved Documents", "Retrieved Pages", "Generated Answer", "Gold Answer",
    "Precision@5", "Recall@5", "Recall@10", "MRR", "nDCG@10",
    "Citation Identity Accuracy", "Citation Gold Precision", "Citation Coverage",
    "RAGAS Context Precision", "RAGAS Context Recall", "RAGAS Answer Relevancy",
    "RAGAS Faithfulness", "RAGAS Correctness", "RAGAS Noise Sensitivity",
    "Latency (ms)", "Input Tokens", "Output Tokens", "Estimated Cost", "Failure Category", "Reviewer Notes",
]


def generate_workbook(report: dict, records: list[dict], output: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    retrieval = {row["id"]: row for row in (report.get("retrieval") or {}).get("rows", [])}
    answers = {row["id"]: row for row in (report.get("answers") or {}).get("rows", [])}
    ragas = {row["id"]: row for row in (report.get("ragas") or {}).get("rows", [])}
    failures = {row["id"]: row for row in (report.get("failure_analysis") or {}).get("rows", [])}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Question Results"
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    for record in records:
        row_r, row_a, row_g = retrieval.get(record["id"], {}), answers.get(record["id"], {}), ragas.get(record["id"], {})
        failure = failures.get(record["id"], {})
        execution = row_a.get("execution_status") or row_r.get("execution_status") or "NOT_RUN"
        row_result = "FAIL" if failure or execution != "COMPLETED" else "PASS"
        sheet.append([
            record["id"], record["question"], record["category"], record["difficulty"], execution, row_result,
            row_a.get("error") or row_r.get("error"), ", ".join(row_r.get("retrieved_document_ids", [])),
            ", ".join(map(str, row_r.get("retrieved_pages", []))), row_a.get("answer"), record["gold_answer"],
            row_r.get("precision_at_5"), row_r.get("recall_at_5"), row_r.get("recall_at_10"),
            row_r.get("reciprocal_rank"), row_r.get("ndcg_at_10"), row_a.get("citation_identity_accuracy"),
            row_a.get("citation_gold_precision"), row_a.get("citation_coverage"),
            row_g.get("context_precision"), row_g.get("context_recall"), row_g.get("answer_relevancy"),
            row_g.get("faithfulness"), row_g.get("answer_correctness"), row_g.get("noise_sensitivity"),
            row_a.get("latency_ms") or row_r.get("latency_ms"), row_a.get("input_tokens"),
            row_a.get("output_tokens"), row_a.get("estimated_cost"), failure.get("failure_category"), "",
        ])
    sheet.freeze_panes, sheet.auto_filter.ref = "A2", sheet.dimensions
    for column, width in {"A": 18, "B": 55, "F": 45, "G": 35, "I": 65, "J": 55, "AC": 35}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value", "Measurement Status"])
    summaries = {
        "Benchmark measurable records": report["benchmark"].get("measurable_count"),
        "Retrieval coverage": report["retrieval"].get("coverage"),
        "Answer coverage": (report.get("answers") or {}).get("coverage"),
        "Recall@5": report["retrieval"].get("recall_at_5"), "Recall@10": report["retrieval"].get("recall_at_10"),
        "MRR": report["retrieval"].get("mrr"), "nDCG@10": report["retrieval"].get("ndcg_at_10"),
        "RAGAS Faithfulness": report["ragas"].get("faithfulness"),
        "Citation identity accuracy": (report.get("answers") or {}).get("citation_identity_accuracy"),
        "Citation coverage": (report.get("answers") or {}).get("citation_coverage"),
    }
    for name, value in summaries.items():
        status = "MEASURED" if value is not None else "NOT_MEASURED"
        summary.append([name, value, status])
    chart = BarChart()
    chart.title = "Measured Retrieval Quality"
    chart.add_data(Reference(summary, min_col=2, min_row=3, max_row=8), titles_from_data=False)
    chart.set_categories(Reference(summary, min_col=1, min_row=3, max_row=8))
    summary.add_chart(chart, "E2")

    gates = workbook.create_sheet("Quality Gates")
    gates.append(["Gate", "Value", "Operator", "Target", "Status", "Passed", "Reason"])
    for name, gate in report["quality_gates"]["gates"].items():
        gates.append([name, gate.get("value"), gate.get("operator"), gate.get("target"), gate.get("status"), gate.get("passed"), gate.get("reason")])
    config = workbook.create_sheet("Run Configuration")
    config.append(["Field", "Value"])
    for key in ("generated_at", "dataset", "evaluation_limit", "pilot_run", "system_versions"):
        config.append([key, json.dumps(report.get(key), ensure_ascii=False)])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
