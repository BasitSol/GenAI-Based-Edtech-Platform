"""Append-only workbook for answers generated from the student interface."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from src.core import ROOT

LIVE_COLUMNS = [
    "Timestamp", "Question", "Generated Answer", "Answer Type", "Category", "Difficulty",
    "Execution Status", "Retrieved Documents", "Retrieved Pages", "Citations", "Citation Valid",
    "Confidence", "Latency (ms)", "Input Tokens", "Output Tokens", "Estimated Cost",
    "RAGAS Status", "RAGAS Context Precision", "RAGAS Context Recall", "RAGAS Faithfulness",
    "RAGAS Answer Relevancy", "RAGAS Answer Correctness", "RAGAS Noise Sensitivity", "Trace ID",
]


def append_live_answer(result: dict, path: Path | None = None) -> Path:
    """Persist one interface answer without calling any additional model."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output = path or (ROOT / "evaluation" / "results" / "live_answers.xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        workbook = load_workbook(output)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Live Answers"
        sheet.append(LIVE_COLUMNS)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True)
        sheet.freeze_panes = "A2"

    profile = result.get("question_understanding") or {}
    debug = result.get("retrieval_debug") or {}
    citations = result.get("citations") or []
    sheet.append([
        datetime.now(timezone.utc).isoformat(), result.get("question", ""), result.get("answer", ""),
        result.get("answer_type"), profile.get("category"), profile.get("difficulty"),
        result.get("execution_status"), ", ".join(sorted({str(x.get("document_id")) for x in result.get("retrieved_chunks", [])})),
        ", ".join(str(x) for x in sorted({x.get("page_start") for x in result.get("retrieved_chunks", []) if x.get("page_start") is not None})),
        "; ".join(f"{x.get('document_id')} p.{x.get('page')}" for x in citations),
        result.get("citation_valid"), result.get("confidence"), result.get("latency_ms"),
        result.get("input_tokens", 0), result.get("output_tokens", 0), result.get("estimated_cost", 0.0),
        "NOT_MEASURED", None, None, None, None, None, None, result.get("trace_id"),
    ])
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"B": 55, "C": 70, "H": 35, "I": 25, "J": 45}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(output)
    return output
